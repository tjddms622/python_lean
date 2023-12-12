import openpyxl

wb = openpyxl.load_workbook('game.xlsx')

ws = wb.active

ws = wb['Sheet1']

print('Total number of rows: ' + str(ws.max_row) + '. And Total number of columns: ' + str(ws.max_column))
print('A1 셀의 값은: ' + ws['A1'].value)

values = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column+1)]

print(values)

data = [ws.cell(row= i, column=2).value for i in range(2,12)]

print(data)

ws['K1'] = 'Sum of Sales'
ws.cell(row=1, column=11, value= 'Sum of Sales')

wb.save('game.xlsx')

row_position = 1

for i in range(1, ws.max_row):

    row_position += 1
    NA_Sales = ws.cell(row=row_position, column=7).value
    EU_Sales = ws.cell(row=row_position, column=8).value
    JP_Sales = ws.cell(row=row_position, column=9).value
    Other_Sales = ws.cell(row=row_position, column=10).value

    total_Sales = (NA_Sales + EU_Sales + JP_Sales + Other_Sales)
    ws.cell(row=row_position, column=11).value = total_Sales

wb.save('game.xlsx')